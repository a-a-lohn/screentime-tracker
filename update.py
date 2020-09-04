from openpyxl import load_workbook, workbook
from re import search, findall
from math import floor
import datetime
# Do this to shorten some method calls (above import is still necessary)
from datetime import date, timedelta
# IMPORTANT: May need to change relative paths below to absolute paths
hist_path = "screentime_tracker/data/HistoryReport.xlsx"
data_path = "screentime_tracker/data/AppData.xlsx"
# Wb is short for workbook
wbh = load_workbook(hist_path)
wbd = load_workbook(data_path)
wbd.active = wbd.get_sheet_by_name("Data")
wsd = wbd.active

def sheetDates(sheet_dates):
    """Create a tupled list of all sheets generated from Samsung app in the form (datetime.date, name)."""    
    sheet_names = (name for name in wbh.sheetnames)
    for name in sheet_names:
        if search("^\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2}$", name):
            # Append date object from sheet name with the sheet name. Date objects are in the form y,m,d
            sheet_dates.append((date(int(name[6:10]), int(name[3:5]), int(name[0:2])), name))

def initTimeInfo():
    """Write column header cells relating to time."""
    wsd['A1'] = 'time_info'
    wsd['$A$2'] = 'Season'
    wsd['$B$2'] = 'DOTW'
    wsd['$C$2'] = 'Date'

def disregardOlderSheets(most_recent_data_date, sheet_dates):
    """Delete sheets from sheet_dates containing data that has already been added."""
    # Convert from datetime.datetime to datetime.date
    most_recent_data_date = date(most_recent_data_date.year, most_recent_data_date.month, most_recent_data_date.day)
    get_later_date = (date[0] for date in sheet_dates)
    i = 0
    while i < len(sheet_dates) and next(get_later_date) <= most_recent_data_date:
        i = i+1
    if i > 0:
        del sheet_dates[:i]
    # Make sure all sheets have been generated on Fridays, if not throw an Exception
    for weeklydate in sheet_dates:
        if weeklydate[0].weekday() != 4:
            raise Exception("Not all sheets were generated on a Friday")
    return sheet_dates

def initAppInfo(sheet_dates):
    """Write app names as column header cells to start."""
    wsd['F1'] = 'app_info'
    # Get the oldest sheet
    wbh.active = wbh.get_sheet_by_name(sheet_dates[0][1])
    wsh = wbh.active
    # Skip over first cell, which is blank and the last few, which have other info
    get_cols = wsh.iter_cols(min_row=2, max_row=wsh.max_row-4, values_only=True)
    # Extract column of apps
    app_names_in_sheet = next(get_cols)
    for app, i in zip(app_names_in_sheet, range(0, len(app_names_in_sheet))):
        wsd.cell(column=i+6, row=2, value=app)

def convertToTime(time_str):
    """Convert a string from a cell in HistoryReport.xlsx to a Datetime.time object."""
    time_str = str(time_str)
    h,m,s = 0,0,0
    # If unit of time (i.e. h/m/s) is not present in cell (i.e. the app was not used long enough or was used
    # for exactly one minute, perhaps), mark as 0 for that unit
    if search("\d{1,2}h", time_str):
        h = int(findall("\d{1,2}h", time_str)[0][:-1])
    if search("\d{1,2}m", time_str):
        m = int(findall("\d{1,2}m", time_str)[0][:-1])
    if search("\d{1,2}s", time_str):
        s = int(findall("\d{1,2}s", time_str)[0][:-1])
    return datetime.time(h,m,s)

def main():
    """Generate an updated graph of screentime usage."""
    sheet_dates = []
    sheetDates(sheet_dates)
    # Sort dates from the earliest to the most recent
    sheet_dates.sort(key=lambda tup: tup[0])

    # Initialize time_info
    if wsd['A1'].value != 'time_info':
        initTimeInfo()
    time_info_cells = wsd['A2:C2']
    
    # Get the most recent date recorded in Data sheet. time_info_cells has all cell data in its first tuple
    most_recent_data_date = wsd.cell(column=time_info_cells[0][2].column, row=wsd.max_row).value
    
    # If Data sheet is not empty, use latest date in sheet to overwrite sheet_dates with only the newest sheets
    if most_recent_data_date != "Date":
        sheet_dates = disregardOlderSheets(most_recent_data_date, sheet_dates)
    
    # Initialize app_info--should only be done if AppData has no data or app name column headers
    if wsd['F1'].value != 'app_info':
        initAppInfo(sheet_dates)
    
    # Create a list with all the app cells in the Data sheet
    app_cells_in_data = []
    for i in range(6, wsd.max_column+1):
        app_cells_in_data.append(wsd.cell(column=i, row=2))
    
    # Create a list of app values (i.e. the string literals of the app names only)
    app_values_in_data = [app.value for app in app_cells_in_data]

    # Keep track of first empty row and column to populate in Data sheet
    new_row_data = wsd.max_row + 1
    new_col_data = wsd.max_column + 1

    # Dict used for the four dates when season changes
    seasons = {date(2020, 3, 21):"Spring", date(2020, 6, 21):"Summer", date(2020, 9, 21):"Fall", date(2020, 12, 21):"Winter"}

    # Dict to convert DOTW (day of the week) number to abbreviation
    days = {0:"M", 1:"Tu", 2:"W", 3:"Th", 4:"F", 5:"Sa", 6:"Su"}

    # Keep track of last app added for use in for loop below
    new_app = False
    for sheet in sheet_dates:
        """Add data to Data sheet week by week by iterating through new sheets in HistoryReport.xlsx which are noted in sheet_dates"""
        app_cells_to_append = []
        app_values_to_append = []

        # 1) Add the time info in Data sheet
        for cell in time_info_cells[0]:
            if cell.value == "Season":
                for i in range(0,7):
                    date2020 = date(2020, sheet[0].month, sheet[0].day)
                    if date2020 in seasons:
                        wsd.cell(column=cell.column, row=new_row_data+i, value=seasons[date2020])
                    else:
                        wsd.cell(column=cell.column, row=new_row_data+i, value=wsd.cell(column=cell.column, row=new_row_data+i-1).value)
            if cell.value == "DOTW":
                for i in range(0,7):
                    date_minus_i = sheet[0] - timedelta(days=6-i)
                    # The DOTW is found by converting the required date (date_minus_i) to a date object,
                    # then using its weekday() attribute to query the days dictionary to return the desired
                    # string abbreviation for the day
                    wsd.cell(column=cell.column, row=new_row_data+i, value=days[date(date_minus_i.year, \
                        date_minus_i.month, date_minus_i.day).weekday()])
            if cell.value == "Date":
                for i in range(0,7):
                    wsd.cell(column=cell.column, row=new_row_data+i, value=sheet[0]-timedelta(days=6-i))

        # 2) Add the app info
        wbh.active = wbh.get_sheet_by_name(sheet[1])
        wsh = wbh.active
        # Skip over first cell, which is blank and the last few, which have other info
        get_cols = wsh.iter_cols(min_row=2, max_row=wsh.max_row-4, values_only=True)
        # Extract column of apps
        app_names_in_sheet = next(get_cols)
        # Generate the app times for each app for each DOTW
        app_row = wsh.iter_rows(min_row=2, max_row=wsh.max_row-4, min_col=2, max_col=8, values_only=True)
        for app in app_names_in_sheet:
            if new_app == True:
                new_col_data += 1
                new_app = False
            if app in app_values_in_data:
                # If app is not new, add data under its existing column header
                app_times = next(app_row)
                # Get values for time spent on app, getting all 7 weekly values for a given app
                for i, time in zip(range(0,7), app_times):
                    time = convertToTime(time)
                    if wsd.cell(column=app_cells_in_data[app_values_in_data.index(app)].column, \
                            row=new_row_data+i).value is None:
                        # Write in the time data under the column header corresponding to the app name,
                        # over 7 rows
                        wsd.cell(column=app_cells_in_data[app_values_in_data.index(app)].column, \
                            row=new_row_data+i, value=time)
                    # The top cell will not be empty if the app name is an already-downloaded duplicate. In this case, find 
                    # the second occurrence of the app and write data over there instead
                    elif app_values_in_data.count(app) > 1:
                        # Start searching through app list for the second occurrence of the app name, starting one position after first occurrence
                        wsd.cell(column=app_cells_in_data[app_values_in_data.index(app, \
                            app_values_in_data.index(app)+1)].column, row=new_row_data+i, value=time)
                    # If this is a new app with the same name as an existing app, add it as a new app column header with the times
                    else:
                        # Write in the app name as a column header (but only once when i==0, not 7 times)
                        if i == 0:
                            wsd.cell(column=new_col_data, row=2, value=app)
                            new_app = True
                        # Write in time values
                        wsd.cell(column=new_col_data, row=new_row_data+i, value=time)

                        app_cells_to_append.append(wsd.cell(column=new_col_data, row=2))
                        app_values_to_append.append(wsd.cell(column=new_col_data, row=2).value)
            # The app must have only been downloaded this past week and does not have the same name as an existing app
            else:
                # Write in the app name as a column header
                wsd.cell(column=new_col_data, row=2, value=app)
                new_app = True
                # Write in time values
                app_times = next(app_row)
                for i, time in zip(range(0,7), app_times):
                    time = convertToTime(time)
                    wsd.cell(column=new_col_data, row=new_row_data+i, value=time)
                app_cells_to_append.append(wsd.cell(column=new_col_data, row=2))
                app_values_to_append.append(wsd.cell(column=new_col_data, row=2).value)
        
        # 3) Add total and running avg column values for each day
        for i in range(0,7):
            total = timedelta()
            for time in wsd.iter_cols(min_row=new_row_data+i, max_row=new_row_data+i, \
                min_col=6, max_col=new_col_data, values_only=True):
                if time[0] is not None:
                    total += timedelta(hours=time[0].hour, minutes=time[0].minute, seconds=time[0].second)
            # Make sure value is of type datetime.timedelta to perform required operations
            if isinstance(wsd.cell(column=5, row=new_row_data+i-1).value, datetime.time):
                last_run_avg = timedelta(hours=wsd.cell(column=5, row=new_row_data+i-1).value.hour, \
                minutes=wsd.cell(column=5, row=new_row_data+i-1).value.minute, \
                seconds=wsd.cell(column=5, row=new_row_data+i-1).value.second)
            else:
                last_run_avg = wsd.cell(column=5, row=new_row_data+i-1).value
            # Take the last running avg value, multiply by num of days it was calculated for, add newest day total
            # and divide by new num of days it is being calculated for (i.e. 1 more day)
            mult_by = new_row_data+i-3
            div_by = new_row_data+i-2
            run_avg = (last_run_avg * mult_by + total) / div_by
            wsd.cell(column=4, row=new_row_data+i, value=total)
            wsd.cell(column=5, row=new_row_data+i, value=run_avg)
        # Add new apps to lists of app cells and app values
        app_cells_in_data = app_cells_in_data + app_cells_to_append
        app_values_in_data = app_values_in_data + app_values_to_append
        new_row_data = new_row_data + 7
    
    # Due to a bug with openpyxl that causes 00:00:00 times to appear as negative values in spreadsheet, cells
    # that had this value must be repopulated with it.
    for row in wsd.iter_rows(min_row=3, min_col=6, max_col=wsd.max_column, max_row=wsd.max_row):
        for cell in row:
            # Convert from datetime.datetime to datetime.time
            if cell.value is not None:
                cell_time = datetime.time(cell.value.hour, cell.value.minute, cell.value.second)
                if cell_time == datetime.time(0,0,0):
                    cell.value = datetime.time(0,0,0)

    wbd.save(data_path)

if __name__ == "__main__":
    main()
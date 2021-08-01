from sys import argv
from openpyxl import load_workbook
from re import search, findall

def convert_to_num_sec(time_str:str) -> int:
    """Convert a cell in HistoryReport.xlsx from a string to an integer indicating the number of seconds."""
    time_str = str(time_str)
    h,m,s = 0,0,0
    if search("\d{1,2}h", time_str):
        h = int(findall("\d{1,2}h", time_str)[0][:-1])
    if search("\d{1,2}m", time_str):
        m = int(findall("\d{1,2}m", time_str)[0][:-1])
    if search("\d{1,2}s", time_str):
        s = int(findall("\d{1,2}s", time_str)[0][:-1])
    return h*3600 + m*60 + s

def convert_all(ws) -> None:
    """Convert all cells in a transposed table from strings to times."""
    for col in ws.iter_cols(min_row=2, min_col=2):
        for cell in col:
            ws.cell(row=cell.row, column=cell.column, value=convert_to_num_sec(cell.value))

def main():
    if len(argv) == 1:
        data_path = f"data/StayFree Export - Total Usage - 5_5_21.xlsx"
    else:
        data_path = f"data/StayFree Export - Total Usage - {argv[1]}_{argv[2]}_21.xlsx"
    wb = load_workbook(data_path)
    wb.active = wb["Usage Time"]
    convert_all(wb.active)
    wb.save(data_path)

if __name__ == "__main__":
    main()

        